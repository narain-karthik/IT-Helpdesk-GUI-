from flask import render_template, request, redirect, url_for, flash, session, abort, make_response, send_file, send_from_directory
from werkzeug.security import generate_password_hash
from werkzeug.utils import secure_filename
from app import app, db
from models import User, Ticket, TicketComment
from forms import LoginForm, TicketForm, UpdateTicketForm, CommentForm, UserRegistrationForm, AssignTicketForm, UserProfileForm
from datetime import datetime
import logging
import os
import socket
import platform
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
import io

# Helper function to check if user is logged in
def is_logged_in():
    return 'user_id' in session

# Helper function to get current user
def get_current_user():
    if is_logged_in():
        return User.query.get(session['user_id'])
    return None

# Helper function to require login
def login_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('user_login'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

# Helper function to require admin
def admin_required(f):
    def decorated_function(*args, **kwargs):
        if not is_logged_in():
            flash('Please log in to access this page.', 'warning')
            return redirect(url_for('admin_login'))
        user = get_current_user()
        if not user or not user.is_admin:
            flash('Admin access required.', 'error')
            return redirect(url_for('index'))
        return f(*args, **kwargs)
    decorated_function.__name__ = f.__name__
    return decorated_function

@app.route('/')
def index():
    """Home page"""
    return render_template('index.html')

@app.route('/user-login', methods=['GET', 'POST'])
def user_login():
    """User login page"""
    if is_logged_in():
        return redirect(url_for('user_dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data) and user.role == 'user':
            session['user_id'] = user.id
            session['is_admin'] = False
            session['role'] = user.role
            
            # Update IP address and system info
            user.ip_address = request.environ.get('HTTP_X_FORWARDED_FOR', request.environ.get('REMOTE_ADDR'))
            try:
                user.system_name = socket.gethostname()
            except:
                user.system_name = 'Unknown'
            db.session.commit()
            
            flash(f'Welcome back, {user.first_name}!', 'success')
            return redirect(url_for('user_dashboard'))
        else:
            flash('Invalid username or password.', 'error')
    
    return render_template('user_login.html', form=form)

@app.route('/admin-login', methods=['GET', 'POST'])
def admin_login():
    """Admin login page"""
    if is_logged_in() and get_current_user().is_admin:
        user = get_current_user()
        if user.is_super_admin:
            return redirect(url_for('super_admin_dashboard'))
        else:
            return redirect(url_for('admin_dashboard'))
    
    form = LoginForm()
    if form.validate_on_submit():
        user = User.query.filter_by(username=form.username.data).first()
        if user and user.check_password(form.password.data) and user.is_admin:
            session['user_id'] = user.id
            session['is_admin'] = True
            session['role'] = user.role
            flash(f'Welcome back, {user.first_name}!', 'success')
            
            if user.is_super_admin:
                return redirect(url_for('super_admin_dashboard'))
            else:
                return redirect(url_for('admin_dashboard'))
        else:
            flash('Invalid admin credentials.', 'error')
    
    return render_template('admin_login.html', form=form)

@app.route('/logout')
def logout():
    """Logout user"""
    session.clear()
    flash('You have been logged out.', 'info')
    return redirect(url_for('index'))

@app.route('/user-dashboard')
@login_required
def user_dashboard():
    """User dashboard showing their tickets"""
    user = get_current_user()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    search_query = request.args.get('search', '')
    
    # Build query
    query = Ticket.query.filter_by(user_id=user.id)
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    return render_template('user_dashboard.html', user=user, tickets=tickets, 
                         status_filter=status_filter, search_query=search_query)

@app.route('/user-profile', methods=['GET', 'POST'])
@login_required
def user_profile():
    """User profile management"""
    user = get_current_user()
    form = UserProfileForm(obj=user)
    
    if form.validate_on_submit():
        user.first_name = form.first_name.data
        user.last_name = form.last_name.data
        user.email = form.email.data
        user.department = form.department.data
        user.system_name = form.system_name.data
        
        # Handle profile image upload
        if 'profile_image' in request.files:
            file = request.files['profile_image']
            if file.filename != '':
                filename = secure_filename(file.filename)
                # Save file logic would go here
                user.profile_image = filename
        
        db.session.commit()
        flash('Profile updated successfully!', 'success')
        return redirect(url_for('user_profile'))
    
    return render_template('user_profile.html', form=form, user=user)

@app.route('/super-admin-dashboard')
@admin_required
def super_admin_dashboard():
    """Super Admin dashboard with full system overview"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    # Get comprehensive statistics
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    
    # Get user statistics
    total_users = User.query.filter_by(role='user').count()
    total_admins = User.query.filter_by(role='admin').count()
    
    # Get recent tickets
    recent_tickets = Ticket.query.order_by(Ticket.created_at.desc()).limit(10).all()
    
    # Get category statistics
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()
    
    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'total_users': total_users,
        'total_admins': total_admins,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets
    }
    
    return render_template('super_admin_dashboard.html', stats=stats, recent_tickets=recent_tickets)

@app.route('/admin-dashboard')
@admin_required
def admin_dashboard():
    """Admin dashboard showing assigned tickets"""
    user = get_current_user()
    
    # Get filter parameters
    status_filter = request.args.get('status', 'all')
    priority_filter = request.args.get('priority', 'all')
    category_filter = request.args.get('category', 'all')
    search_query = request.args.get('search', '')
    
    # Build query - only show tickets assigned to this admin
    query = Ticket.query.filter_by(assigned_to=user.id)
    
    if status_filter != 'all':
        query = query.filter_by(status=status_filter)
    
    if priority_filter != 'all':
        query = query.filter_by(priority=priority_filter)
    
    if category_filter != 'all':
        query = query.filter_by(category=category_filter)
    
    if search_query:
        query = query.filter(Ticket.title.contains(search_query))
    
    tickets = query.order_by(Ticket.created_at.desc()).all()
    
    # Get statistics for assigned tickets
    total_assigned = Ticket.query.filter_by(assigned_to=user.id).count()
    open_assigned = Ticket.query.filter_by(assigned_to=user.id, status='Open').count()
    in_progress_assigned = Ticket.query.filter_by(assigned_to=user.id, status='In Progress').count()
    resolved_assigned = Ticket.query.filter_by(assigned_to=user.id, status='Resolved').count()
    
    stats = {
        'total': total_assigned,
        'open': open_assigned,
        'in_progress': in_progress_assigned,
        'resolved': resolved_assigned
    }
    
    return render_template('admin_dashboard.html', tickets=tickets, stats=stats,
                         status_filter=status_filter, priority_filter=priority_filter,
                         category_filter=category_filter, search_query=search_query, admin_user=user)

@app.route('/create-ticket', methods=['GET', 'POST'])
@login_required
def create_ticket():
    """Create a new ticket"""
    form = TicketForm()
    user = get_current_user()
    
    if form.validate_on_submit():
        # Capture current session IP and system info for this specific ticket
        # Get real client IP address (handles proxy forwarding)
        current_ip = request.environ.get('HTTP_X_FORWARDED_FOR', 
                                       request.environ.get('HTTP_X_REAL_IP', 
                                       request.environ.get('REMOTE_ADDR')))
        
        # Get system name - prioritize user input or preserve detected system name
        current_system_name = None
        if form.system_name.data and form.system_name.data.strip():
            # Use the actual system name entered by user or detected by client
            current_system_name = form.system_name.data.strip()
        else:
            # Fallback to current user's system name if available, otherwise use generic detection
            if user.system_name and user.system_name.strip():
                current_system_name = user.system_name.strip()
            else:
                # Only use generic detection as last resort
                user_agent = request.headers.get('User-Agent', '').lower()
                if 'windows' in user_agent:
                    current_system_name = 'Windows System'
                elif 'mac os x' in user_agent or 'macos' in user_agent:
                    current_system_name = 'macOS System'
                elif 'linux' in user_agent:
                    current_system_name = 'Linux System'
                elif 'android' in user_agent:
                    current_system_name = 'Android Device'
                elif 'iphone' in user_agent or 'ipad' in user_agent:
                    current_system_name = 'iOS Device'
                else:
                    current_system_name = f'Unknown System ({request.remote_addr})'
        
        # Update user's profile with latest info (optional)
        user.ip_address = current_ip
        user.system_name = current_system_name
        
        # Handle image upload
        image_filename = None
        if form.image.data and form.image.data.filename:
            file = form.image.data
            if file.filename != '':
                filename = secure_filename(file.filename)
                if filename:  # Make sure secure_filename didn't return empty string
                    # Create unique filename with timestamp
                    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S_')
                    image_filename = timestamp + filename
                    
                    # Ensure upload directory exists
                    upload_dir = 'static/uploads'
                    os.makedirs(upload_dir, exist_ok=True)
                    
                    filepath = os.path.join(upload_dir, image_filename)
                    try:
                        file.save(filepath)
                    except Exception as e:
                        flash('Error uploading image. Ticket created without image.', 'warning')
                        image_filename = None
        
        ticket = Ticket(
            title=form.title.data,
            description=form.description.data,
            category=form.category.data,
            priority=form.priority.data,
            user_id=user.id,
            user_name=user.full_name,
            user_ip_address=current_ip,
            user_system_name=current_system_name,
            image_filename=image_filename
        )
        db.session.add(ticket)
        db.session.commit()
        
        flash(f'Ticket {ticket.ticket_number} created successfully!', 'success')
        return redirect(url_for('user_dashboard'))
    
    return render_template('create_ticket.html', form=form)

@app.route('/ticket/<int:ticket_id>')
@login_required
def view_ticket(ticket_id):
    """View ticket details"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can view this ticket
    if not user.is_admin and ticket.user_id != user.id:
        abort(403)
    
    form = CommentForm()
    assign_form = AssignTicketForm() if user.is_admin else None
    
    return render_template('view_ticket.html', ticket=ticket, form=form, 
                         assign_form=assign_form, user=user)

@app.route('/ticket/<int:ticket_id>/comment', methods=['POST'])
@login_required
def add_comment(ticket_id):
    """Add comment to ticket"""
    ticket = Ticket.query.get_or_404(ticket_id)
    user = get_current_user()
    
    # Check if user can comment on this ticket
    if not user.is_admin and ticket.user_id != user.id:
        abort(403)
    
    form = CommentForm()
    if form.validate_on_submit():
        comment = TicketComment(
            ticket_id=ticket_id,
            user_id=user.id,
            comment=form.comment.data
        )
        db.session.add(comment)
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('Comment added successfully!', 'success')
    
    return redirect(url_for('view_ticket', ticket_id=ticket_id))

@app.route('/ticket/<int:ticket_id>/edit', methods=['GET', 'POST'])
@admin_required
def edit_ticket(ticket_id):
    """Edit ticket (admin only)"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = UpdateTicketForm(obj=ticket)
    
    if form.validate_on_submit():
        ticket.title = form.title.data
        ticket.description = form.description.data
        ticket.category = form.category.data
        ticket.priority = form.priority.data
        
        old_status = ticket.status
        ticket.status = form.status.data
        
        # Set resolved_at if status changed to Resolved
        if old_status != 'Resolved' and ticket.status == 'Resolved':
            ticket.resolved_at = datetime.utcnow()
        elif ticket.status != 'Resolved':
            ticket.resolved_at = None
        
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        flash('Ticket updated successfully!', 'success')
        return redirect(url_for('view_ticket', ticket_id=ticket_id))
    
    return render_template('edit_ticket.html', form=form, ticket=ticket)

@app.route('/ticket/<int:ticket_id>/assign', methods=['POST'])
@admin_required
def assign_ticket(ticket_id):
    """Assign ticket to admin"""
    ticket = Ticket.query.get_or_404(ticket_id)
    form = AssignTicketForm()
    
    if form.validate_on_submit():
        ticket.assigned_to = form.assigned_to.data
        if ticket.status == 'Open':
            ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        assignee = User.query.get(form.assigned_to.data)
        flash(f'Ticket assigned to {assignee.full_name}!', 'success')
    
    return redirect(url_for('view_ticket', ticket_id=ticket_id))

@app.route('/manage-users')
@admin_required
def manage_users():
    """Super Admin user management"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    users = User.query.all()
    return render_template('manage_users.html', users=users)

@app.route('/create-user', methods=['GET', 'POST'])
@admin_required
def create_user():
    """Create new user (Super Admin only)"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    form = UserRegistrationForm()
    if form.validate_on_submit():
        new_user = User(
            username=form.username.data,
            email=form.email.data,
            first_name=form.first_name.data,
            last_name=form.last_name.data,
            department=form.department.data,
            role=form.role.data,
            is_admin=(form.role.data in ['admin', 'super_admin'])
        )
        new_user.set_password(form.password.data)
        db.session.add(new_user)
        db.session.commit()
        
        flash(f'User {new_user.username} created successfully!', 'success')
        return redirect(url_for('manage_users'))
    
    return render_template('create_user.html', form=form)

@app.route('/assign-work/<int:ticket_id>', methods=['GET', 'POST'])
@admin_required
def assign_work(ticket_id):
    """Super Admin assigns work to specific admins based on category"""
    user = get_current_user()
    if not user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    # Get appropriate admins based on ticket category
    if ticket.category in ['Hardware']:
        admins = User.query.filter_by(role='admin', department='IT Hardware').all()
    elif ticket.category in ['Software']:
        admins = User.query.filter_by(role='admin', department='IT Software').all()
    else:
        admins = User.query.filter_by(role='admin').all()
    
    form = AssignTicketForm()
    form.assigned_to.choices = [(admin.id, f"{admin.full_name} ({admin.department})") for admin in admins]
    
    if form.validate_on_submit():
        ticket.assigned_to = form.assigned_to.data
        ticket.status = 'In Progress'
        ticket.updated_at = datetime.utcnow()
        db.session.commit()
        
        assignee = User.query.get(form.assigned_to.data)
        flash(f'Work assigned to {assignee.full_name}!', 'success')
        return redirect(url_for('super_admin_dashboard'))
    
    return render_template('assign_work.html', form=form, ticket=ticket, admins=admins)

def create_default_admin():
    """Create default admin and super admin users if none exists"""
    try:
        super_admin = User.query.filter_by(role='super_admin').first()
        if not super_admin:
            # Create Super Admin
            super_admin_user = User(
                username='superadmin',
                email='superadmin@gtnengineering.com',
                first_name='Super',
                last_name='Administrator',
                department='IT',
                role='super_admin',
                is_admin=True
            )
            super_admin_user.set_password('super123')
            db.session.add(super_admin_user)
            db.session.commit()
            
            # Create Hardware Admins
            hardware_admins = [
                {'username': 'yuvaraj', 'first_name': 'Yuvaraj', 'last_name': 'Admin'},
                {'username': 'jayachandran', 'first_name': 'Jayachandran', 'last_name': 'Admin'},
                {'username': 'narainkarthik', 'first_name': 'Narain', 'last_name': 'Karthik'}
            ]
            
            for admin_data in hardware_admins:
                admin_user = User(
                    username=admin_data['username'],
                    email=f"{admin_data['username']}@gtnengineering.com",
                    first_name=admin_data['first_name'],
                    last_name=admin_data['last_name'],
                    department='IT Hardware',
                    role='admin',
                    is_admin=True
                )
                admin_user.set_password('admin123')
                db.session.add(admin_user)
                db.session.commit()
            
            # Create Software Admins
            software_admins = [
                {'username': 'sathish', 'first_name': 'Sathish', 'last_name': 'SAP Admin'},
                {'username': 'lakshmiprabha', 'first_name': 'Lakshmi', 'last_name': 'Prabha'}
            ]
            
            for admin_data in software_admins:
                admin_user = User(
                    username=admin_data['username'],
                    email=f"{admin_data['username']}@gtnengineering.com",
                    first_name=admin_data['first_name'],
                    last_name=admin_data['last_name'],
                    department='IT Software',
                    role='admin',
                    is_admin=True
                )
                admin_user.set_password('admin123')
                db.session.add(admin_user)
                db.session.commit()
            
            # Create a test user
            test_user = User(
                username='testuser',
                email='user@gtnengineering.com',
                first_name='Test',
                last_name='User',
                department='Engineering',
                role='user',
                is_admin=False
            )
            test_user.set_password('test123')
            db.session.add(test_user)
            db.session.commit()
            
            logging.info("Default super admin, admins and test user created")
    except Exception as e:
        logging.error(f"Error creating default users: {e}")
        db.session.rollback()

@app.route('/reports-dashboard')
@admin_required
def reports_dashboard():
    """Reports Dashboard with visual analytics (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    # Get comprehensive statistics
    total_tickets = Ticket.query.count()
    open_tickets = Ticket.query.filter_by(status='Open').count()
    in_progress_tickets = Ticket.query.filter_by(status='In Progress').count()
    resolved_tickets = Ticket.query.filter_by(status='Resolved').count()
    closed_tickets = Ticket.query.filter_by(status='Closed').count()
    
    # Category breakdown
    hardware_tickets = Ticket.query.filter_by(category='Hardware').count()
    software_tickets = Ticket.query.filter_by(category='Software').count()
    network_tickets = Ticket.query.filter_by(category='Network').count()
    other_tickets = Ticket.query.filter_by(category='Other').count()
    
    # Priority breakdown  
    critical_tickets = Ticket.query.filter_by(priority='Critical').count()
    high_tickets = Ticket.query.filter_by(priority='High').count()
    medium_tickets = Ticket.query.filter_by(priority='Medium').count()
    low_tickets = Ticket.query.filter_by(priority='Low').count()
    
    # Get all tickets for detailed table
    all_tickets = Ticket.query.order_by(Ticket.created_at.desc()).all()
    
    stats = {
        'total_tickets': total_tickets,
        'open_tickets': open_tickets,
        'in_progress_tickets': in_progress_tickets,
        'resolved_tickets': resolved_tickets,
        'closed_tickets': closed_tickets,
        'hardware_tickets': hardware_tickets,
        'software_tickets': software_tickets,
        'network_tickets': network_tickets,
        'other_tickets': other_tickets,
        'critical_tickets': critical_tickets,
        'high_tickets': high_tickets,
        'medium_tickets': medium_tickets,
        'low_tickets': low_tickets
    }
    
    # Prepare chart data for JavaScript
    chart_data = {
        'category': [hardware_tickets, software_tickets, network_tickets, other_tickets],
        'priority': [critical_tickets, high_tickets, medium_tickets, low_tickets],
        'status': [open_tickets, in_progress_tickets, resolved_tickets, closed_tickets]
    }
    
    return render_template('reports_dashboard.html', stats=stats, tickets=all_tickets, chart_data=chart_data)

@app.route('/edit-assignment/<int:ticket_id>', methods=['GET', 'POST'])
@admin_required
def edit_assignment(ticket_id):
    """Edit ticket assignment (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    ticket = Ticket.query.get_or_404(ticket_id)
    
    if request.method == 'POST':
        assigned_to = request.form.get('assigned_to')
        if assigned_to:
            assigned_to = int(assigned_to) if assigned_to != '0' else None
        else:
            assigned_to = None
            
        ticket.assigned_to = assigned_to
        ticket.updated_at = datetime.utcnow()
        
        try:
            db.session.commit()
            assignee_name = User.query.get(assigned_to).full_name if assigned_to else 'Unassigned'
            flash(f'Ticket {ticket.ticket_number} has been assigned to {assignee_name}.', 'success')
            return redirect(url_for('super_admin_dashboard'))
        except Exception as e:
            db.session.rollback()
            flash('Error updating assignment. Please try again.', 'error')
            logging.error(f"Error updating ticket assignment: {e}")
    
    # Get all admin users for assignment dropdown
    admin_users = User.query.filter_by(is_admin=True).all()
    
    return render_template('edit_assignment.html', ticket=ticket, admin_users=admin_users)

@app.route('/view-image/<filename>')
@admin_required
def view_image(filename):
    """View uploaded ticket image (Admin and Super Admin only)"""
    try:
        return send_from_directory('static/uploads', filename)
    except FileNotFoundError:
        abort(404)

@app.route('/download-excel-report')
@admin_required
def download_excel_report():
    """Download Excel report of all tickets (Super Admin only)"""
    current_user = get_current_user()
    if not current_user.is_super_admin:
        flash('Super Admin access required.', 'error')
        return redirect(url_for('index'))
    
    try:
        # Create a new workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Tickets Report"
        
        # Define headers
        headers = [
            'Ticket ID', 'Title', 'Description', 'Category', 'Priority', 'Status',
            'Created By', 'User Email', 'User Department', 'System Name', 'IP Address',
            'Assigned To', 'Created Date', 'Updated Date', 'Resolved Date'
        ]
        
        # Style headers
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # Add headers to worksheet
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
        
        # Get all tickets with related data
        tickets = Ticket.query.join(User, Ticket.user_id == User.id).all()
        
        # Add ticket data
        for row, ticket in enumerate(tickets, 2):
            assignee_name = ticket.assignee.full_name if ticket.assignee else 'Unassigned'
            
            data = [
                ticket.ticket_number,
                ticket.title,
                ticket.description,
                ticket.category,
                ticket.priority,
                ticket.status,
                ticket.user_name,
                ticket.user.email,
                ticket.user.department or 'N/A',
                ticket.user_system_name or 'N/A',
                ticket.user_ip_address or 'N/A',
                assignee_name,
                ticket.created_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.created_at else 'N/A',
                ticket.updated_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.updated_at else 'N/A',
                ticket.resolved_at.strftime('%Y-%m-%d %H:%M:%S') if ticket.resolved_at else 'N/A'
            ]
            
            for col, value in enumerate(data, 1):
                ws.cell(row=row, column=col, value=value)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to memory
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Generate filename with timestamp
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f'GTN_Helpdesk_Report_{timestamp}.xlsx'
        
        # Create response
        response = make_response(output.getvalue())
        response.headers['Content-Type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        response.headers['Content-Disposition'] = f'attachment; filename={filename}'
        
        return response
        
    except Exception as e:
        logging.error(f"Error generating Excel report: {e}")
        flash('Error generating report. Please try again.', 'error')
        return redirect(url_for('reports_dashboard'))

# Initialize default admin on first import
with app.app_context():
    create_default_admin()

# Error handlers
@app.errorhandler(404)
def not_found_error(error):
    return render_template('404.html'), 404

@app.errorhandler(403)
def forbidden_error(error):
    return render_template('403.html'), 403

@app.errorhandler(500)
def internal_error(error):
    db.session.rollback()
    return render_template('500.html'), 500
